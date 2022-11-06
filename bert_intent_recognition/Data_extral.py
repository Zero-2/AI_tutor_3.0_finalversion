import csv
import openpyxl
import random
import pandas as pd
import os
import xlwt

seed = 233
random.seed(seed)
os.environ['PYTHONHASHSEED'] = str(seed)

def gen_sample_base_template(eneity):
    # definition
    definition_explain_qwds = ["what is ", "Could you tell me ", "Can i ask ", "Can you describe ", "May i ask ",
                               "what can you tell me ", "How to describe "]
    definition_qwds = ["the meaning of ", "the definition of ", "the description of ", "the portrait of ",
                       "the picture of "]

    # developer
    developer_explain_qwds = ['Who ', 'which person ', 'which one ']
    developer_qwds = ['is the funder of ', 'set up the ', 'is the designer ', 'is the inventor ', "is the developer ",
                      "is the creator ", ]

    # different
    different_explain_qwds = ['What is ', 'How is ', 'Is there a ', 'Is there any ', 'Is the ', 'Are there a ',
                              'Are there any ', 'Are the ']
    different_qwds = ['the difference ', 'the contrast ', 'the gap ', 'the separation ']
    different_connectives_qwds = ['between ', 'compare ', 'from ']

    # drawback
    drawback_explain_qwds = ['What are ', 'What is ', "Could you tell me ", "Can i ask ", "Can you describe ",
                             "May i ask ", "what can you tell me ", "How to describe ", 'Are there any ']
    drawback_qwds = ['the drawbacks of ', ' the disadvantages of ', 'the downsides of ', 'the shortcomings of ',
                     'the cons of ']

    # example
    example_explain_qwds = ["Could you tell me ", "Can i ask ", "May i ask ", "Can you give me ", 'Are there any ',
                            "Can you provide ", " Can you show me ", "Can you offer me ", "Can you point out "]
    example_qwds = ["the example of ", "the illustrate of ", "the instance of ", "the sample of ", "the case of "]

    # method
    method_explain_qwds = ["What do we need to ", "How to ", "what is the requirement to ",
                           "which method can we used to "]
    method_qwds = ["achieve the ", "use the ", "realize the ", "accomplish the ", "implement ", "carry out ", ]

    # reason
    reason_explain_qwds = ["what is the reason of ", "Can you explain the reason of ",
                           "Could you tell me the reason of ", "Can you give me the reason about ","what is the factor to cause "]

    # represent
    represent_explain_qwds = ["what is "]
    represent_qwds = ["represented as ","figured as ","traced as "]

    #use_to
    use_to_explain_qwds = ["what is "]
    use_to_qwds = ["used for ", "directed at ", "used in ","used on "]

    use_to_explain_qwds2 = ["where ","which area that ","which area"]
    use_to_qwds2 = ["can be used for ", "can be directed at ", "can be used in ", "can be used on "]

    greet_qwds = ['Hello ! ', 'excuse me ! ', 'Hi ! ', 'Could you please tell me that ', 'I want to know that ',]

    label_list = [line.strip() for line in open('data/label', 'r', encoding='utf8')]
    label2id = {label: idx for idx, label in enumerate(label_list)}

    eneity_list = eneity
    n = len(eneity_list) - 1

    data = []
    data_train = []

    # 问定义
    template = "{greet}{explain}{definition}{eneity} ?"
    for i in range(150):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        explain = definition_explain_qwds[random.randint(0, len(definition_explain_qwds) - 1)]
        definition = definition_qwds[random.randint(0, len(definition_qwds) - 1)]
        eneity = eneity_list[random.randint(0, n)][0]
        text = template.format(greet=greet, explain=explain, definition=definition, eneity=eneity, )
        text=text.lower()
        data.append([text, 'definition', label2id['definition']])
        data_train.append([0,0,0,0,text,eneity])

    # developer
    template = "{greet}{explain}{developer}{eneity} ?"
    for i in range(300):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        explain = developer_explain_qwds[random.randint(0, len(developer_explain_qwds) - 1)]
        eneity = eneity_list[random.randint(0, n)][0]
        developer = developer_qwds[random.randint(0, len(developer_qwds) - 1)]
        text = template.format(greet=greet, explain=explain, developer=developer, eneity=eneity)
        text = text.lower()
        data.append([text, 'developer', label2id['developer']])
        data_train.append([0, 0, 0, 0, text, eneity])

    # different
    template = "{greet}{explain}{difference}{Connectives}{eneity1} and {eneity2} ?"
    for i in range(300):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        eneity1 = eneity_list[random.randint(0, n)][0]
        eneity2 = eneity_list[random.randint(0, n)][0]
        explain = different_explain_qwds[random.randint(0, len(different_explain_qwds) - 1)]
        difference = different_qwds[random.randint(0, len(different_qwds) - 1)]
        Connectives = different_connectives_qwds[random.randint(0, len(different_connectives_qwds) - 1)]
        text = template.format(greet=greet, explain=explain, difference=difference, Connectives=Connectives,
                               eneity1=eneity1, eneity2=eneity2)
        text = text.lower()
        data.append([text, 'different', label2id['different']])
        data_train.append([0, 0, 0, 0, text, eneity])

    # drawback
    template = "{greet}{explain}{drawback}{eneity} ?"
    for i in range(300):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        explain = drawback_explain_qwds[random.randint(0, len(drawback_explain_qwds) - 1)]
        eneity = eneity_list[random.randint(0, n)][0]
        drawback = drawback_qwds[random.randint(0, len(drawback_qwds) - 1)]
        text = template.format(greet=greet, explain=explain, drawback=drawback, eneity=eneity)
        text = text.lower()
        data.append([text, 'drawback', label2id['drawback']])
        data_train.append([0, 0, 0, 0, text, eneity])

    # example
    template = "{greet}{explain}{example}{eneity} ?"
    for i in range(300):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        explain = example_explain_qwds[random.randint(0, len(example_explain_qwds) - 1)]
        eneity = eneity_list[random.randint(0, n)][0]
        example = example_qwds[random.randint(0, len(example_qwds) - 1)]
        text = template.format(greet=greet, explain=explain, example=example, eneity=eneity)
        text = text.lower()
        data.append([text, 'example', label2id['example']])
        data_train.append([0, 0, 0, 0, text, eneity])

    # has_part

    # method
    template = "{greet}{explain}{method}{eneity} ?"
    for i in range(300):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        explain = method_explain_qwds[random.randint(0, len(method_explain_qwds) - 1)]
        eneity = eneity_list[random.randint(0, n)][0]
        method = method_qwds[random.randint(0, len(method_qwds) - 1)]
        text = template.format(greet=greet, explain=explain, method=method, eneity=eneity)
        text = text.lower()
        data.append([text, 'method', label2id['method']])
        data_train.append([0, 0, 0, 0, text, eneity])

    # reason
    template = "{greet}{explain}{eneity} ?"
    for i in range(300):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        explain = reason_explain_qwds[random.randint(0, len(reason_explain_qwds) - 1)]
        eneity = eneity_list[random.randint(0, n)][0]
        text = template.format(greet=greet, explain=explain, eneity=eneity)
        text = text.lower()
        data.append([text, 'reason', label2id['reason']])
        data_train.append([0, 0, 0, 0, text, eneity])

    # represent
    template = "{greet}{explain}{eneity}{represent} ?"
    for i in range(300):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        explain = represent_explain_qwds[random.randint(0, len(represent_explain_qwds) - 1)]
        eneity = eneity_list[random.randint(0, n)][0]
        represent = represent_qwds[random.randint(0, len(represent_qwds) - 1)]
        text = template.format(greet=greet, explain=explain, eneity=eneity, represent=represent)
        text = text.lower()
        data.append([text, 'represent', label2id['represent']])
        data_train.append([0, 0, 0, 0, text, eneity])

    # the_same_thing
    # template = "{greet}{explain}{eneity}{the_same_thing} ?"
    # for i in range(300):
    #     greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
    #     explain = represent_explain_qwds[random.randint(0, len(represent_explain_qwds) - 1)]
    #     eneity = eneity_list[random.randint(0, n)][0]
    #     the_same_thing = represent_qwds[random.randint(0, len(represent_qwds) - 1)]
    #     text = template.format(greet=greet, explain=explain, eneity=eneity, the_same_thing=the_same_thing)
    #     text = text.lower()
    #     data.append([text, 'the_same_thing', label2id['the_same_thing']])

    # use_to
    template = "{greet}{explain}{eneity}{use_to} ?"
    for i in range(100):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        explain = use_to_explain_qwds[random.randint(0, len(use_to_explain_qwds) - 1)]
        eneity = eneity_list[random.randint(0, n)][0]
        use_to = use_to_qwds[random.randint(0, len(use_to_qwds) - 1)]
        text = template.format(greet=greet, explain=explain, eneity=eneity, use_to=use_to)
        text = text.lower()
        data.append([text, 'use_to', label2id['use_to']])
        data_train.append([0, 0, 0, 0, text, eneity])
    for i in range(200):
        greet = greet_qwds[random.randint(0, len(greet_qwds) - 1)]
        explain = use_to_explain_qwds2[random.randint(0, len(use_to_explain_qwds2) - 1)]
        eneity = eneity_list[random.randint(0, n)][0]
        use_to = use_to_qwds2[random.randint(0, len(use_to_qwds2) - 1)]
        text = template.format(greet=greet, explain=explain, eneity=eneity, use_to=use_to)
        text = text.lower()
        data.append([text, 'use_to', label2id['use_to']])
        data_train.append([0, 0, 0, 0, text, eneity])


    return data, data_train


def deal_data(path, sheet_name):

    label_list = [line.strip() for line in open('data/label', 'r', encoding='utf8')]
    id_label = {label: index for index, label in enumerate(label_list)}

    data = []
    Total_eneity = []
    chapters = [1,2,3,4,5,8,9,10,11,12,13,14,15]

    for i in chapters:
        full_path = path + str(i)+'.xlsx'
        workbook = openpyxl.load_workbook(full_path)
        workbook_sheet = workbook[sheet_name]

        for row in range(2, workbook_sheet.max_row):
            row_stack = []
            text = workbook_sheet.cell(row, 5).value
            text = text.lower()
            label_class = workbook_sheet.cell(row, 7).value
            if label_class not in label_list:
                continue
            label = id_label.get(label_class)
            eneity = workbook_sheet.cell(row, 6).value.lower()
            eneity = [eneity]

            if text == None or label_class == None:
                continue

            if eneity != None and eneity not in Total_eneity:
                Total_eneity.append(eneity)

            row_stack.append(text)
            row_stack.append(label_class)
            row_stack.append(label)
            data.append(row_stack)

    return data, Total_eneity

def creat_dataset(data):
    random.shuffle(data)

    with open("data/train.csv", "w+", encoding="utf-8-sig", newline='') as f:
        write = csv.writer(f)
        write.writerow(["text","label_class","label"])
        for item in range(0,int(len(data)*0.7)):
            write.writerow(data[item])

    with open("data/test.csv", "w+", encoding="utf-8-sig", newline='') as f:
        write = csv.writer(f)
        write.writerow(["text","label_class","label"])
        for item in range(int(len(data)*0.7)+1,len(data)-1):
            write.writerow(data[item])

def creat_eneityset(eneities):

    with open("data/eneities.csv", "w+", encoding="utf-8-sig", newline='') as f:
        write = csv.writer(f)
        write.writerows(eneities)

def creat_training_data(data):
    book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheet = book.add_sheet('Q&A', cell_overwrite_ok=True)
    col = ('1','2','3','4','Question','Intention entity')
    for i in range(0, 6):
        sheet.write(0, i, col[i])

    for i in range(len(data)):
        data_row = data[i]
        for j in range(0, 6):
            sheet.write(i + 1, j, data_row[j])

    savepath = './data/Bilstm_crf_data.xls'
    book.save(savepath)

if __name__ == '__main__':

    if not os.path.exists("./checkpoint"):
        os.mkdir('./checkpoint')

    path = "../Concepts/"
    sheet_name = "Q&A"

    data1, eneity = deal_data(path, sheet_name)
    creat_eneityset(eneity)

    data2, data_train = gen_sample_base_template(eneity)
    creat_training_data(data_train)
    data = data1 + data2
    creat_dataset(data)